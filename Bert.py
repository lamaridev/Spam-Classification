import pandas as pd
import openpyxl
import torch
from torch.utils.data import DataLoader, TensorDataset
from sklearn.model_selection import train_test_split
from transformers import BertForSequenceClassification, AdamW , BertTokenizer


def nettoyer_texte(texte):
    # Convert text to lowercase
    texte = texte.lower()
    texte = texte.replace('"', '')  
    return texte

def load_data(filepath):
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        # Extract questions and answers
        category = [nettoyer_texte(str(sheet['A' + str(i)].value)) for i in range(2, sheet.max_row + 1)]
        context = [nettoyer_texte(str(sheet['B' + str(i)].value)) for i in range(2, sheet.max_row + 1)]

        # Assuming you map spam to 1 and ham to 0
        category = [1 if cat == 'spam' else 0 for cat in category]

        return category, context

    except Exception as e:
        print(f"Error reading XLSX file: {e}")
        return []

def create_tensors(inputs):
    tokenizer = BertTokenizer.from_pretrained('bert-base-uncased')
    input_tensors = tokenizer(inputs, padding=True, truncation=True, return_tensors="pt")
    return input_tensors

def create_dataloader(input_ids, attention_mask, labels, batch_size, shuffle=True):
    # Combine the input tensors and labels into a TensorDataset
    dataset = TensorDataset(input_ids, attention_mask, labels)

    # Create DataLoader
    dataloader = DataLoader(dataset, batch_size=batch_size, shuffle=shuffle)

    return dataloader

def classification_bert(train_dataloader, epochs, learning_rate, device="cuda" if torch.cuda.is_available() else "cpu"):

    # 1. Initialize the BERT model for sequence classification
    model = BertForSequenceClassification.from_pretrained("bert-base-uncased",
        num_labels=2,  # For binary classification (spam/ham)
        output_attentions=False,
        output_hidden_states=False
    )
    model.to(device)  # Move model to the specified device

    # 2. Define loss function, optimizer, and scheduler
    optimizer = AdamW(model.parameters(), lr=learning_rate)
    loss_fn = torch.nn.CrossEntropyLoss().to(device)

    # 3. Training loop
    for epoch in range(epochs):
        total_loss = 0
        model.train()

        for step, batch in enumerate(train_dataloader):
            input_ids = batch[0].to(device)
            attention_mask = batch[1].to(device)
            labels = batch[2].to(device)

            # Clear previous gradients
            model.zero_grad()

            # Forward pass
            outputs = model(input_ids, attention_mask=attention_mask, labels=labels)
            loss = outputs[0]
            total_loss += loss.item()

            # Backward pass
            loss.backward()

            # Update weights
            optimizer.step()

            # Print the status every 5 batches
            if (step + 1) % 10 == 0:
                average_loss = total_loss / (step + 1)
                print(f"Epoch [{epoch + 1}/{epochs}] | Batch [{step + 1}/{len(train_dataloader)}] | Loss: {average_loss:.4f}")

    model.save_pretrained("/content/model")
    return model  # Return the fine-tuned model



# Step 1: Load data
categories, contexts = load_data('/content/spam.xlsx')

# Split the data
contexts_train, contexts_test, categories_train, categories_test = train_test_split(contexts, categories, test_size=0.2, random_state=42)

# Step 2: Tokenize text and convert to tensors for TRAINING data
encoded_inputs_train = create_tensors(contexts_train)
input_ids_train = encoded_inputs_train['input_ids']
attention_mask_train = encoded_inputs_train['attention_mask']
labels_train = torch.tensor(categories_train)

# For TESTING data
encoded_inputs_test = create_tensors(contexts_test)
input_ids_test = encoded_inputs_test['input_ids']
attention_mask_test = encoded_inputs_test['attention_mask']
labels_test = torch.tensor(categories_test)

# Step 3: Create dataloader for TRAINING data

train_dataloader = create_dataloader(input_ids_train, attention_mask_train, labels_train, batch_size=32)

# For TESTING data
test_dataloader = create_dataloader(input_ids_test, attention_mask_test, labels_test, batch_size=32, shuffle=False)


trained_model = classification_bert(train_dataloader, epochs=15, learning_rate=1e-3)